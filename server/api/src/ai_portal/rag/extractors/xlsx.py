"""XLSX extractor via :mod:`openpyxl`.

Each sheet becomes one :class:`TableBlock` whose ``meta.sheet`` carries
the sheet name. Empty sheets are skipped.
"""
from __future__ import annotations

from io import BytesIO
from typing import Any

from ai_portal.rag.extractors.protocol import (
    Block,
    ExtractedDocument,
    TableBlock,
)


class XlsxExtractor:
    name = "xlsx"
    mime_types = {
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "application/vnd.ms-excel",
    }

    def supports(self, mime: str) -> bool:
        return mime in self.mime_types

    async def extract(self, data: bytes, meta: dict[str, Any]) -> ExtractedDocument:
        from openpyxl import load_workbook  # type: ignore

        wb = load_workbook(BytesIO(data), data_only=True, read_only=True)
        blocks: list[Block] = []
        text_parts: list[str] = []
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows: list[list[str]] = []
            for row in ws.iter_rows(values_only=True):
                cells = ["" if c is None else str(c) for c in row]
                if any(c.strip() for c in cells):
                    rows.append(cells)
            if not rows:
                continue
            blocks.append(TableBlock(rows=rows, sheet=sheet_name))
            text_parts.append(
                "\n".join("\t".join(r) for r in rows)
            )
        return ExtractedDocument(
            text="\n\n".join(text_parts),
            blocks=blocks,
            meta={**meta, "sheet_count": len(wb.sheetnames)},
        )


__all__ = ["XlsxExtractor"]
